# apollo_pipeline.py — Deployment Ready Version

# ─────────────────────────
# 🔧 Imports & Setup
# ─────────────────────────

from apify_client import ApifyClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import asyncio
import aiohttp
import nest_asyncio
nest_asyncio.apply()
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import SequenceMatcher


# === CONFIGURATION ===

# Constants to be set via input()
EXCEL_FILE = "apollo_scraped_data.xlsx"
EMAIL_COL_NAME = "Personal Email"
LINKEDIN_COL_NAME = "Personal LinkedIn"

# === STEP FUNCTIONS ===
# === STEP 1: Scrape Apollo Data ===
def step1_scrape_apollo():
    print("▶ Running Step 1: Scraping Apollo data...")

    client = ApifyClient(APIFY_API_TOKEN)
    actor_input = {
        "searchUrl": APOLLO_SEARCH_URL,
        "count": TOTAL_RECORDS
    }

    run = client.actor("supreme_coder/apollo-scraper").call(run_input=actor_input)
    dataset_items = client.dataset(run["defaultDatasetId"]).list_items().items

    extracted_data = []
    for item in dataset_items:
        extracted_data.append({
            "First Name": item.get("firstName"),
            "Last Name": item.get("lastName"),
            "Company Name": item.get("company", {}).get("companyName"),
            "Company Website": item.get("company", {}).get("websiteUrl"),
            "Company Domain": item.get("company", {}).get("mainDomain"),
            "Company Address": item.get("company", {}).get("fullAddress"),
            "Personal Email": item.get("emailAddress"),
            "Personal LinkedIn": item.get("linkedInProfileUrl"),
            "Company LinkedIn": item.get("company", {}).get("linkedInProfileUrl"),
            "Job Title": item.get("headline"),
            "Company Employee Count": item.get("company", {}).get("employeeEstimate"),
            "Company Country": item.get("company", {}).get("countryName"),
        })

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    headers = list(extracted_data[0].keys())
    ws1.append(headers)

    for data in extracted_data:
        ws1.append([data.get(header) for header in headers])

    wb.save(EXCEL_FILE)
    print(f"✅ Step 1 complete. {len(extracted_data)} contacts saved to {EXCEL_FILE}.")


# === STEP 2: Verify Emails with Bounceban ===
def step2_verify_emails():
    print("▶ Running Step 2: Verifying emails with Bounceban...")

    # Load Excel workbook and sheet
    wb = load_workbook(EXCEL_FILE)
    ws1 = wb["Sheet1"]

    # Get headers and index of the email column
    headers = [cell.value for cell in ws1[1]]
    email_col_index = headers.index(EMAIL_COL_NAME) + 1

    # Collect emails and row numbers
    email_rows = []
    for row in ws1.iter_rows(min_row=2, values_only=False):
        email_cell = row[email_col_index - 1]
        email = email_cell.value
        if email:
            email_rows.append((email, email_cell.row))

    # === Asynchronous verification ===
    async def verify_email(session, email):
        url = "https://api.bounceban.com/v1/verify/single"
        params = {"email": email}
        headers = {"Authorization": BOUNCEBAN_API_KEY}
        try:
            async with session.get(url, params=params, headers=headers) as resp:
                data = await resp.json()
                return email, data.get("state", "unknown")
        except Exception:
            return email, "error"

    async def verify_all_emails(email_list):
        results = {}
        async with aiohttp.ClientSession() as session:
            tasks = [verify_email(session, email) for email in email_list]
            responses = await asyncio.gather(*tasks)
            for email, status in responses:
                results[email] = status
        return results

    async def wrapper():
        emails = [email for email, _ in email_rows]
        results = await verify_all_emails(emails)
    
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        undeliverable_count = 0  # ✅ Initialize the counter BEFORE the loop
    
        # Process results
        for email, row_num in email_rows:
            state = results.get(email, "unknown")
    
            # Manually handle Apollo placeholder emails
            if "email_not_unlocked" in email:
                state = "undeliverable"
    
            # Print log for every email and status
            print(f"Email: {email}, Status: {state}")
    
            if state == "undeliverable":
                cell = ws1.cell(row=row_num, column=email_col_index)
                cell.fill = red_fill
                cell.value = None  # Just remove the email, keep the rest of the row
                undeliverable_count += 1
    
        wb.save(EXCEL_FILE)
        print(f"✅ Step 2 complete. {undeliverable_count} undeliverable emails removed and marked in red.")

    asyncio.get_event_loop().run_until_complete(wrapper())
    

# === STEP 3: Extract Personal Informaiton Through LinkedIn Scrapper ===

def step3_scrape_linkedin_profiles():
    print("▶ Running Step 3: Scraping LinkedIn profiles...")

    wb = load_workbook(EXCEL_FILE)
    ws1 = wb["Sheet1"]

    # Clean Sheet2 if it exists
    if "Sheet2" in wb.sheetnames:
        del wb["Sheet2"]
    ws2 = wb.create_sheet("Sheet2")

    headers = [cell.value for cell in ws1[1]]
    linkedin_col_index = headers.index(LINKEDIN_COL_NAME) + 1

    linkedin_urls = []
    row_mapping = []

    for row in ws1.iter_rows(min_row=2, values_only=False):
        cell = row[linkedin_col_index - 1]
        url = cell.value
        if url:
            linkedin_urls.append(url)
            row_mapping.append(cell.row)

    client = ApifyClient(APIFY_API_TOKEN)
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    scraped_profiles = []

    for url, row_num in zip(linkedin_urls, row_mapping):
        try:
            input_payload = {"profileUrls": [url]}
            run = client.actor("dev_fusion/Linkedin-Profile-Scraper").call(run_input=input_payload)
            result = client.dataset(run["defaultDatasetId"]).list_items().items

            if not result or len(result) == 0:
                raise Exception("No data returned")

            profile = result[0]

            scraped_profiles.append({
                "First Name": profile.get("firstName", ""),
                "Last Name": profile.get("lastName", ""),
                "Headline": profile.get("headline", ""),
                "Email": profile.get("email", ""),
                "Job Title": profile.get("jobTitle", ""),
                "Company Name": profile.get("companyName", ""),
                "Company Address": profile.get("addressWithCountry", "")
            })

            print(f"✅ Scraped: {url}")

        except Exception as e:
            print(f"❌ Failed to scrape: {url} — {str(e)}")
            scraped_profiles.append({
                "First Name": "",
                "Last Name": "",
                "Headline": "",
                "Email": "",
                "Job Title": "",
                "Company Name": "",
                "Company Address": ""
            })
            cell = ws1.cell(row=row_num, column=linkedin_col_index)
            cell.fill = red_fill

    # Write header row to Sheet2
    ws2.append([
        "First Name", "Last Name", "Headline", "Email", 
        "Job Title", "Company Name", "Company Address"
    ])

    for profile in scraped_profiles:
        ws2.append([
            profile["First Name"],
            profile["Last Name"],
            profile["Headline"],
            profile["Email"],
            profile["Job Title"],
            profile["Company Name"],
            profile["Company Address"]
        ])

    wb.save(EXCEL_FILE)
    print(f"✅ Step 3 complete. Scraped {len(scraped_profiles)} profiles to Sheet2.")
    
# === STEP 4: Fill Missing emails which are present in either sheets ===

def step4_fill_missing_emails():
    print("▶ Step 4: Filling missing emails between Sheet1 and Sheet2...")

    wb = load_workbook(EXCEL_FILE)
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]

    # Read headers
    sheet1_headers = [cell.value for cell in ws1[1]]
    sheet2_headers = [cell.value for cell in ws2[1]]

    # Get column indices
    s1_fname_idx = sheet1_headers.index("First Name") + 1
    s1_lname_idx = sheet1_headers.index("Last Name") + 1
    s1_email_idx = sheet1_headers.index("Personal Email") + 1

    s2_fname_idx = sheet2_headers.index("First Name") + 1
    s2_lname_idx = sheet2_headers.index("Last Name") + 1
    s2_email_idx = sheet2_headers.index("Email") + 1

    # Create maps: {(first + last): email}
    sheet1_map = {}
    for row in ws1.iter_rows(min_row=2, values_only=False):
        key = (row[s1_fname_idx - 1].value or "").strip().lower() + "_" + (row[s1_lname_idx - 1].value or "").strip().lower()
        email = row[s1_email_idx - 1].value
        sheet1_map[key] = {"row": row, "email": email}

    sheet2_map = {}
    for row in ws2.iter_rows(min_row=2, values_only=False):
        key = (row[s2_fname_idx - 1].value or "").strip().lower() + "_" + (row[s2_lname_idx - 1].value or "").strip().lower()
        email = row[s2_email_idx - 1].value
        sheet2_map[key] = {"row": row, "email": email}

    updated_s1 = 0
    updated_s2 = 0

    # Copy emails into Sheet1 where missing
    for key in sheet1_map:
        s1_email = sheet1_map[key]["email"]
        s2_email = sheet2_map.get(key, {}).get("email")

        if not s1_email and s2_email:
            cell = sheet1_map[key]["row"][s1_email_idx - 1]
            cell.value = s2_email
            updated_s1 += 1

    # Copy emails into Sheet2 where missing
    for key in sheet2_map:
        s2_email = sheet2_map[key]["email"]
        s1_email = sheet1_map.get(key, {}).get("email")

        if not s2_email and s1_email:
            cell = sheet2_map[key]["row"][s2_email_idx - 1]
            cell.value = s1_email
            updated_s2 += 1

    wb.save(EXCEL_FILE)
    print(f"✅ Step 4 complete. Emails filled: {updated_s1} in Sheet1, {updated_s2} in Sheet2.")


# === STEP 5: Company Name Comparision in Both Sheets ===

def step5_compare_company_names():
    print("▶ Step 5: Comparing company names between Sheet1 and Sheet2...")

    wb = load_workbook(EXCEL_FILE)
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]

    # Create or replace Sheet3
    if "Sheet3" in wb.sheetnames:
        del wb["Sheet3"]
    ws3 = wb.create_sheet("Sheet3")

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    def is_similar(name1, name2, threshold=0.8):
        if not name1 or not name2:
            return False
        return SequenceMatcher(None, name1.lower(), name2.lower()).ratio() >= threshold

    # Create key maps with correct columns
    sheet1_map = {}
    for row in ws1.iter_rows(min_row=2):
        first = (row[0].value or "").strip().lower()
        last = (row[1].value or "").strip().lower()
        name_key = f"{first}_{last}"
        company_name = row[2].value  # ✅ Correct: Column C = "Company Name"
        sheet1_map[name_key] = company_name

    sheet2_map = {}
    for row in ws2.iter_rows(min_row=2):
        first = (row[0].value or "").strip().lower()
        last = (row[1].value or "").strip().lower()
        name_key = f"{first}_{last}"
        company_name = row[5].value  # Column F = "Company Name"
        sheet2_map[name_key] = company_name

    # Write header
    ws3.append(["First Name", "Last Name", "Company Name S1", "Company Name S2", "Company Name Match"])

    # Compare and write rows
    for name_key in set(sheet1_map.keys()).union(sheet2_map.keys()):
        first, last = name_key.split("_")
        s1_name = sheet1_map.get(name_key, "")
        s2_name = sheet2_map.get(name_key, "")

        # Auto-fill if one side is empty
        if not s1_name and s2_name:
            s1_name = s2_name
        elif not s2_name and s1_name:
            s2_name = s1_name

        match_status = "Match" if is_similar(str(s1_name), str(s2_name)) else "Mismatch"
        row_data = [first.title(), last.title(), s1_name, s2_name, match_status]
        ws3.append(row_data)

        if match_status == "Mismatch":
            for col in range(1, 6):
                ws3.cell(row=ws3.max_row, column=col).fill = red_fill

    wb.save(EXCEL_FILE)
    print("✅ Step 5 complete. Results saved to Sheet3.")



# === STEP 6: Fixing Last Names===

def step6_fix_last_names():
    print("▶ Step 6: Fixing incomplete last names in Sheet1 using Sheet2...")

    wb = load_workbook(EXCEL_FILE)
    ws1 = wb["Sheet1"]
    ws2 = wb["Sheet2"]

    # Read all names from Sheet2 into a dictionary
    sheet2_names = {}
    for row in ws2.iter_rows(min_row=2):
        first = (row[0].value or "").strip().lower()
        last = (row[1].value or "").strip()
        key = first
        if key:
            sheet2_names[key] = last

    updated_count = 0

    for row in ws1.iter_rows(min_row=2):
        first = (row[0].value or "").strip().lower()
        last = (row[1].value or "").strip()

        if first and (not last or len(last) <= 2 or last.endswith(".")):
            # If incomplete, try to find a better match
            corrected_last = sheet2_names.get(first)
            if corrected_last and corrected_last.lower() != last.lower():
                row[1].value = corrected_last
                updated_count += 1

    wb.save(EXCEL_FILE)
    print(f"✅ Step 6 complete. Updated {updated_count} last names in Sheet1.")



# === MASTER RUNNER ===
def run_all_steps():
    step1_scrape_apollo()
    step2_verify_emails()
    step3_scrape_linkedin_profiles()
    step4_fill_missing_emails()
    step5_compare_company_names()
    step6_fix_last_names()
    
    
# === CLI MENU ===
def menu():
    global APIFY_API_TOKEN, BOUNCEBAN_API_KEY, APOLLO_SEARCH_URL, TOTAL_RECORDS

    print("\n🔐 Enter your API keys and input data:")
    APIFY_API_TOKEN = input("Enter Apify API Token: ").strip()
    BOUNCEBAN_API_KEY = input("Enter Bounceban API Key: ").strip()
    APOLLO_SEARCH_URL = input("Paste Apollo People Tab URL: ").strip()
    TOTAL_RECORDS = int(input("Enter number of records to scrape: ").strip())

    while True:
        print("\n📋 Menu")
        print("1 - Step 1: Scrape Apollo")
        print("2 - Step 2: Verify Emails")
        print("3 - Step 3: Scrape LinkedIn")
        print("4 - Step 4: Fill Missing Emails")
        print("5 - Step 5: Compare Company Names")
        print("6 - Step 6: Fix Last Names")
        print("7 - Run All Steps")
        print("0 - Exit")

        choice = input("Choose an option (0-7): ")
        if choice == "1": step1_scrape_apollo()
        elif choice == "2": step2_verify_emails()
        elif choice == "3": step3_scrape_linkedin_profiles()
        elif choice == "4": step4_fill_missing_emails()
        elif choice == "5": step5_compare_company_names()
        elif choice == "6": step6_fix_last_names()
        elif choice == "7": run_all_steps()
        elif choice == "0": print("👋 Exiting..."); break
        else: print("❌ Invalid option. Try again.")

# === RUN ===
if __name__ == "__main__":
    menu()
